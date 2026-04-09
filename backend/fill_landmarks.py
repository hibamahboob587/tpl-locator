import os
import time
import re
import json
import math
import requests
import openpyxl
from pathlib import Path
from dotenv import load_dotenv

# ───────────────── CONFIG ─────────────────

EXCEL_PATH = Path("CardTagReadings_Final.xlsx")
SHEETS = ["STP3","STP4","STP5","STP6","STP7","STP8","STP9","STP10","STP11","STP12"]

COORD_COL    = 3
LAT_COL      = 3
LNG_COL      = 4
LANDMARK_COL = 5

HEADER_ROW   = 2
DATA_START   = 3

MAX_RETRIES  = 3
TIMEOUT      = 30
API_DELAY    = 0.5

# Path to Customlocations.json — looks in frontend/src/utils/ relative to this script
SCRIPT_DIR = Path(__file__).resolve().parent
POI_PATHS = [
    SCRIPT_DIR / "Customlocations.json",
    SCRIPT_DIR / ".." / "frontend" / "src" / "utils" / "Customlocations.json",
    SCRIPT_DIR / ".." / "src" / "utils" / "Customlocations.json",
    SCRIPT_DIR / "frontend" / "src" / "utils" / "Customlocations.json",
]

POI_MAX_DISTANCE = 100  # meters — match POI within this radius

# ─────────────── LOAD ENV TOKEN ───────────────

def load_env(path=".env"):
    env = {}
    try:
        with open(path) as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    k, v = line.split("=", 1)
                    env[k.strip()] = v.strip().strip('"').strip("'")
    except FileNotFoundError:
        pass
    return env

env = load_env()
load_dotenv()
MAPBOX_TOKEN = env.get("MAPBOX_TOKEN") or os.environ.get("MAPBOX_TOKEN")

if not MAPBOX_TOKEN:
    raise SystemExit("❌ MAPBOX_TOKEN not found in .env file")

# ─────────────── LOCAL POI ───────────────

def load_pois():
    for p in POI_PATHS:
        resolved = p.resolve()
        if resolved.exists():
            try:
                with open(resolved, encoding="utf-8") as f:
                    data = json.load(f)
                print(f"📍 Loaded {len(data)} POIs from {resolved}\n")
                return data
            except Exception as e:
                print(f"⚠ Failed to load {resolved}: {e}")
    print("⚠ Customlocations.json not found in any expected path, skipping local POI matching\n")
    return []

POI_LIST = load_pois()

def haversine_meters(lat1, lng1, lat2, lng2):
    R = 6371000
    rlat1, rlat2 = math.radians(lat1), math.radians(lat2)
    dlat = math.radians(lat2 - lat1)
    dlng = math.radians(lng2 - lng1)
    a = math.sin(dlat / 2) ** 2 + math.cos(rlat1) * math.cos(rlat2) * math.sin(dlng / 2) ** 2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

def find_nearest_poi(lat, lng):
    best_name = None
    best_dist = POI_MAX_DISTANCE
    for poi in POI_LIST:
        dist = haversine_meters(lat, lng, poi["lat"], poi["lng"])
        if dist < best_dist:
            best_dist = dist
            best_name = poi["name"]
    return best_name

# ─────────────── CACHE ───────────────

_cache = {}

# ─────────────── UTILITIES ───────────────

def parse_coords(coord_str):
    try:
        parts = [p for p in re.split(r"[,\s]+", str(coord_str).strip()) if p]
        return float(parts[0]), float(parts[1])
    except Exception:
        return None, None


def extract_area_context(feature):
    context_parts = []
    for ctx in feature.get("context", []):
        ctx_id  = ctx.get("id", "")
        ctx_text = ctx.get("text", "")
        if any(ctx_id.startswith(t) for t in ("neighborhood", "locality", "place", "district")):
            context_parts.append(ctx_text)
    return ", ".join(context_parts[:2])


def mapbox_request(url, attempt=0):
    try:
        r = requests.get(url, timeout=TIMEOUT)
        r.raise_for_status()
        return r.json().get("features", [])
    except Exception as e:
        if attempt < MAX_RETRIES - 1:
            wait = 3 * (attempt + 1)
            print(f"    ⚠ Timeout, retrying in {wait}s... (attempt {attempt + 2}/{MAX_RETRIES})")
            time.sleep(wait)
            return mapbox_request(url, attempt + 1)
        else:
            raise e


def reverse_geocode(lat, lng):
    cache_key = f"{lat:.5f},{lng:.5f}"
    if cache_key in _cache:
        return _cache[cache_key]

    base_url = (
        f"https://api.mapbox.com/geocoding/v5/mapbox.places/"
        f"{lng},{lat}.json"
        f"?access_token={MAPBOX_TOKEN}"
        f"&language=en"
    )

    result = ""

    try:
        features = mapbox_request(base_url + "&types=address")
        if features:
            best = features[0]
            addr = best.get("text", "")
            area = extract_area_context(best)
            result = f"{addr}, {area}" if area else addr

        if not result:
            features = mapbox_request(base_url + "&types=neighborhood")
            if features:
                result = features[0].get("place_name", "")

        if not result:
            features = mapbox_request(base_url + "&types=locality")
            if features:
                result = features[0].get("place_name", "")

        if not result:
            features = mapbox_request(base_url + "&types=place")
            if features:
                result = features[0].get("place_name", "")

        if not result:
            result = f"{lat:.6f}, {lng:.6f}"

    except Exception as e:
        print(f"  ❌ Failed after {MAX_RETRIES} attempts at ({lat}, {lng}): {e}")
        result = f"{lat:.6f}, {lng:.6f}"

    parts = [p.strip() for p in result.split(",")]
    parts = [p for p in parts if p.lower() not in ("pakistan",)]
    clean = ", ".join(parts[:3])

    _cache[cache_key] = clean
    time.sleep(API_DELAY)
    return clean


# ─────────────── RESOLVE LANDMARK ───────────────

def resolve_landmark(lat, lng):
    poi = find_nearest_poi(lat, lng)
    if poi:
        return poi, True
    return reverse_geocode(lat, lng), False


# ─────────────── COLUMN SETUP ───────────────

def replace_coord_with_lat_lng(ws):
    if ws.cell(HEADER_ROW, LAT_COL).value == "Latitude":
        print("    ↳ Lat/Lng columns already exist, skipping.")
        return

    coord_data = {}
    for row in range(DATA_START, ws.max_row + 1):
        coord_str = ws.cell(row, COORD_COL).value
        if coord_str:
            coord_data[row] = parse_coords(coord_str)

    ws.insert_cols(LNG_COL, amount=1)

    ws.cell(HEADER_ROW, LAT_COL).value = "Latitude"
    ws.cell(HEADER_ROW, LNG_COL).value = "Longitude"

    for row, (lat, lng) in coord_data.items():
        if lat is not None:
            ws.cell(row, LAT_COL).value = lat
            ws.cell(row, LNG_COL).value = lng


# ─────────────── MAIN ───────────────

def main():
    print(f"📂 Loading {EXCEL_PATH.name}...\n")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    total_updated = 0
    poi_matches   = 0
    api_calls     = 0

    for sheet_name in SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠ Sheet '{sheet_name}' not found, skipping.")
            continue

        ws = wb[sheet_name]
        print(f"📋 Processing {sheet_name}...")

        replace_coord_with_lat_lng(ws)

        updated = 0
        for row in range(DATA_START, ws.max_row + 1):
            lat = ws.cell(row, LAT_COL).value
            lng = ws.cell(row, LNG_COL).value

            if lat is None or lng is None:
                continue

            try:
                lat = float(lat)
                lng = float(lng)
            except (TypeError, ValueError):
                continue

            existing = ws.cell(row, LANDMARK_COL).value
            if existing and str(existing).strip():
                continue

            landmark, was_poi = resolve_landmark(lat, lng)
            ws.cell(row, LANDMARK_COL).value = landmark
            updated += 1

            if was_poi:
                poi_matches += 1
            else:
                api_calls += 1

            if updated % 25 == 0:
                print(f"    {updated} rows updated...")

        print(f"  ✅ {updated} landmarks updated\n")
        total_updated += updated

    print("💾 Saving file...")
    wb.save(EXCEL_PATH)
    print(f"✅ Done — {total_updated} total landmarks updated across all sheets.")
    print(f"   📍 {poi_matches} from local POI  |  🌐 {api_calls} from Mapbox API")


if __name__ == "__main__":
    main()